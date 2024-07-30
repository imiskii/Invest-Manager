##
# Author: Michal Ľaš
# Date: 16.07.2024

from typing import Any, Dict, Generator

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table 
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import rows_from_range
from collections import defaultdict


TableRow = Dict[str, Any]


def iter_table_rows(ws:Worksheet, tb:Table) -> Generator[TableRow, None, None]:
    """ 
    Iterate over rows from a table with headers (row as dictionary)
    Code from Stack Overflow: https://stackoverflow.com/questions/65895165/how-to-iterate-over-all-rows-in-an-excel-table-using-openpyxl
    Author:  Jean-Francois T.
    Date:  Jan 26, 2021 
    """
    def get_row_values(ws:Worksheet,row_cell_ref:tuple):
        return [ws[c].value for c in row_cell_ref]

    iter_rows = rows_from_range(tb.ref)
    headers = get_row_values(ws,next(iter_rows))

    for row_cells in iter_rows:
        yield {h:v for h,v in zip(headers, get_row_values(ws,row_cells))}



class DataLoaderError(Exception):

    def __init__(self, message) -> None:
        self.message = message
        super().__init__(message)

    def __str__(self) -> str:
        return f"DataLoaderError: {self.message}"



class DataLoader:
    """ Load data from .xlsx file and process necessary values """

    _rec_data = defaultdict(list) # dictionary of assets names and list of its records
    _assets_data = defaultdict(TableRow) # dictionary of assets names and the info about asset
    _category_data = defaultdict(float)
    _results_data = defaultdict(TableRow) # dictionary of assets names and the performance of assets

    def __init__(self) -> None:
        pass


    def _read_rec_data(self, record_sheet_name='Records', record_table_name='rec_tab') -> None:
        """ Read records data from .xlsx """
        ws = self._wb[record_sheet_name]
        rec_tb = ws.tables[record_table_name]
        for row in iter_table_rows(ws, rec_tb):
            self._rec_data[row['TICKER']].append(row)


    def _read_assets_data(self, assets_sheet_name='Assets', assets_table_name='tic_tab') -> None:
        """ Read assets data from .xlsx """
        ws = self._wb[assets_sheet_name]
        assets_tb = ws.tables[assets_table_name]
        for row in iter_table_rows(ws, assets_tb):
            self._assets_data[row['TICKER']] = row


    def _read_category_data(self, category_sheet_name='Assets', category_table_name='category_tab') -> None:
        ws = self._wb[category_sheet_name]
        category_tb = ws.tables[category_table_name]
        for row in iter_table_rows(ws, category_tb):
            self._category_data[row['Category']] = row['Goal']


    def _read_results_data(self, results_sheet_name='Results', results_table_name='results_tab') -> None:
        """ Read results data from .xlsx """
        ws = self._wb[results_sheet_name]
        results_tb = ws.tables[results_table_name]
        for row in iter_table_rows(ws, results_tb):
            self._results_data[row['TICKER']] = row


    def read_portfolio_excel(self, xlsx_path:str) -> None:
        """
        Read the given excel file with portfolio data.
        """
        self._xlsx_path = xlsx_path

        try:
            self._wb = load_workbook(self._xlsx_path, data_only=True)
            self._read_rec_data()
            self._read_assets_data()
            self._read_results_data()
            self._read_category_data()
        except FileNotFoundError as e:
            raise DataLoaderError(f"Invalid filepath: {self._xlsx_path}. {e}")
        except Exception as e:
            raise DataLoaderError(f"Error while loading MS Excel portfolio file: {e}")
            

    def get_ticker_data(self, ticker: str) -> dict|None:
        """
        Return data from .xlsx file by ticker in dictionary format:
        'records': list of records
        'info': information about asset
        'results': performance of asset

        It returns 'None' value if given `ticker` is unknow.
        """

        ticker = ticker.upper()

        if ticker not in self._assets_data:
            return None
        else:
            return {'records': self._rec_data[ticker], 'info': self._assets_data[ticker], 'results': self._results_data[ticker]}


    def get_categories(self) -> dict:
        """
        Return data about assets categories.
        """
        return self._category_data
    

    def get_all_tickers(self) -> list[str]:
        return list(self._assets_data.keys())


# END OF FILE #